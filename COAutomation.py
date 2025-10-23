import sys
from flask import Flask, render_template, make_response,request, jsonify, send_file
from openpyxl import Workbook, styles,load_workbook
from openpyxl.styles import Side,PatternFill
import pandas as pd
import os
import io
from io import BytesIO
import os.path
from flask import send_file
import logging
from openpyxl.styles import Font,Border
from openpyxl.utils import get_column_letter
import pdfplumber
import re
from openpyxl import workbook, load_workbook
from FINALTEST import second_bp
from datetime import datetime
import warnings
warnings.filterwarnings("ignore", message="CropBox missing from /Page")

# Define a bold border  
bold_border = styles.Border(left=styles.Side(border_style='thin', color='000000'),
                right=styles.Side(border_style='thin', color='000000'),
                top=styles.Side(border_style='thin', color='000000'),
                
                bottom=styles.Side(border_style='thin', color='000000'))

# set font
font_bold = Font(name="Times New Roman", size=10, bold=True)
font = Font(name="Times New Roman", size=10)

qpCount = 0
coCount = 6
maxRows = 75
template_static_name = "template.xlsx"
template_dynamic_name = ""

app = Flask(__name__)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s: %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler(sys.stdout)
    ]
)

logging.basicConfig(level=logging.INFO)

@app.before_request
def log_request_info():
    logging.info(f"Accessed route: {request.path}")

app.register_blueprint(second_bp)

# Function to check if the file type is allowed
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'pdf'}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload1', methods=['POST'])
def upload():
    if 'pdf_files' not in request.files:
        return jsonify({"message": "No files part in the request"}), 400

    uploaded_files = request.files.getlist('pdf_files')
    if not uploaded_files or all(file.filename == '' for file in uploaded_files):
        return jsonify({"message": "No selected files"}), 400

    upload_folder = 'uploads'
    os.makedirs(upload_folder, exist_ok=True)

    saved_files = []
    for file in uploaded_files:
        if file and file.filename.endswith('.pdf'):
            file_path = os.path.join(upload_folder, file.filename)
            file.save(file_path)
            saved_files.append(file_path)

    # Pass the saved files to the generate_excel function
    generate_excel(saved_files)

    return jsonify({
    "success": True,
    "message": f"Successfully processed {len(saved_files)} file(s).",
    "download_url": f"/download/{template_dynamic_name}"
    }), 200

@app.route("/generate_excel", methods=["POST"])
def generate_excel(pdf_paths):
    
    # Initialize empty lists for storing data for each set of questions, marks, and COs
    ct1_QNos = []
    ct1_Marks = []
    ct1_COs = []
    ct2_QNos = []
    ct2_Marks = []
    ct2_COs = []
    ct3_QNos = []
    ct3_Marks = []
    ct3_COs = []

    # Determine the number of question papers based on the number of pdf_paths
    ct1_grouping = {}
    ct2_grouping = {}
    ct3_grouping = {}
    
    # Determine the number of question papers based on the number of pdf_paths
    global qpCount
    qpCount = len(pdf_paths)

    # Extract details from the first question paper PDF
    ct1_grouping = extract_details_from_pdf(pdf_paths[0], ct1_QNos, ct1_Marks, ct1_COs)
    
    # Update the Question No Choices
    ct1_QNos = update_QuestionNo_Choices(ct1_QNos)

    # If there are more than one question papers, process the second one
    if qpCount > 1:
        ct2_grouping = extract_details_from_pdf(pdf_paths[1], ct2_QNos, ct2_Marks, ct2_COs)
        # Update the Question No Choices
        ct2_QNos = update_QuestionNo_Choices(ct2_QNos)

    # If there are more than two question papers, process the third one 
    if qpCount > 2:
        ct3_grouping = extract_details_from_pdf(pdf_paths[2], ct3_QNos, ct3_Marks, ct3_COs)
        # Update the Question No Choices
        ct3_QNos = update_QuestionNo_Choices(ct3_QNos)

    # Create a new workbook
    workbook = load_workbook(os.path.join(os.path.join(os.getcwd(),"template"),template_static_name))

    worksheet = workbook["CT1-3"]
    if (worksheet is None):
        worksheet = workbook.create_sheet("CT1-3")
    workbook.active = workbook.index(worksheet)

    # Calculate the length of question numbers for each question paper  
    ct1=len(ct1_QNos)
    ct2=len(ct2_QNos)
    ct3=len(ct3_QNos)

    # Generate various rows in the worksheet (headers, data, and formulas)
    generate_first_row(worksheet,ct1,ct2,ct3)
    coColumns = generate_second_row(worksheet,ct1_grouping,ct2_grouping,ct3_grouping)
    generate_third_row(worksheet,ct1,ct2,ct3)
    generate_fourth_row(worksheet,ct1_Marks,ct2_Marks,ct3_Marks,ct1_grouping,ct2_grouping,ct3_grouping)
    generate_fifth_row(worksheet,ct1,ct2,ct3)
    generate_sixth_row(worksheet,ct1_QNos, ct2_QNos, ct3_QNos,ct1_grouping, ct2_grouping, ct3_grouping)
    generate_Formulas(worksheet,ct1,ct2,ct3, coColumns)
    apply_styles(worksheet)

    # Define the static folder for downloads
    static_folder = os.path.join(os.getcwd(), "download")
    os.makedirs(static_folder, exist_ok=True)  # Ensure the folder exists
    # Save file to static folder
    file_path = os.path.join(static_folder, template_dynamic_name)

    # Save to disk before sending
    workbook.save(file_path)  

    # Clean up uploaded files (optional - remove if you want to keep them)
    for pdf_path in pdf_paths:
        if os.path.exists(pdf_path):
            os.remove(pdf_path)

    return jsonify({
        "success": True,
        "message": "Excel file generated successfully.",
        "download_url": "/download/"+template_dynamic_name  # Adjust the download URL as needed
    }), 200

def apply_styles(worksheet):
    # Define colors
    light_green_fill = styles.PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    orange_fill = styles.PatternFill(start_color="FFC300", end_color="FFC300", fill_type="solid")
    yellow_fill = styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    grey_fill = styles.PatternFill(start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")

    # Apply formatting (borders, fills, font styles, and alignment)
    for row_num,row in enumerate(worksheet.iter_rows(min_row=1, max_row=maxRows), 1):
            for cell in row:
                cell.border=bold_border  # Apply border to all cells
            # Apply conditional fill color based on row number
                match row_num:
                    case 1:
                        cell.fill = light_green_fill
                    case 2:
                        cell.fill = orange_fill
                    case 4:
                        cell.fill = yellow_fill
                    case 5:
                        if cell.column != 1:
                            cell.fill = grey_fill
                    case 72:
                        cell.fill = yellow_fill
                # Bold font for certain rows
                if row_num <= 6 or (row_num >= 71 and cell.column < 4):
                    cell.font = font_bold
                else:
                    cell.font = font
                    
                # Center alignment for all columns except the first
                if cell.column != 3:
                    cell.alignment = styles.Alignment(horizontal='center', vertical='center')

def extract_details_from_pdf(pdf_path, question_numbers, marks, co_lists):
    coGrouping = {}

    # Define a regex pattern to match question numbers (e.g., "1", "2", etc.)
    flag=0

    global template_dynamic_name

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            # Extract all text from the page
            text = page.extract_text()

            if text: 
                text = text.replace('&', '"&"')
                 # Ensure the page has text
                # Split the text into lines
                lines = text.split("\n")

                prevLine=''
                for line in lines:
                    if flag== 0 and not (template_dynamic_name.endswith('.xlsx')) and line.find('CYCLE TEST')!=-1:
                        prevLine=line
                        continue
                    if flag== 0 and not (template_dynamic_name.endswith('.xlsx')) and prevLine != "":
                        template_dynamic_name=line.strip()
                        prevLine=""
                        continue
                    if flag== 0 and not (template_dynamic_name.endswith('.xlsx')) and line.find('For')!=-1:
                        template_dynamic_name+="_" + line.replace('(For ','').replace(' / ','_').replace(".","_").replace(': ','_').replace(')','').strip()
                        continue
                    if flag==0 and line.find("Part")!=-1:
                        flag=1
                        continue
                    if flag==1:
                        match = re.search(r".\d \d \d \d$",line.strip()) 
                        question_no= re.match(r"^(\d{2})|\d",line.strip()) 
                        if question_no and match:
                            # Add the matched number to the list
                            try:
                                qnum=question_no.group().strip()
                                question_numbers.append("Q"+qnum)
                                marks.append(int(match.group().strip().split()[0]))
                                co_lists.append(match.group().strip().split()[1])

                            except ValueError:
                                pass  # Skip if the match isn't a valid integer
 
    if '&' in template_dynamic_name:
        template_dynamic_name = re.sub(r'[<>:"/\\|?*]', '', template_dynamic_name.replace('&', 'And'))                        
    if not (template_dynamic_name.endswith('.xlsx')):
        template_dynamic_name=f"{template_dynamic_name}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    
    # Get the CO grouping indices
    for i in range(1,coCount+1):
        co = get_matching_value_indices(co_lists,str(i))
        if len(co) > 0:
            coGrouping[i] = co
    return coGrouping

def update_QuestionNo_Choices(qNos):
    updated_list = qNos.copy()  # Work with a copy of the list to avoid modifying the original
    
    for i in range(len(qNos) - 1):
        if qNos[i] == qNos[i + 1]:
            updated_list[i] = f"{qNos[i]}A"
            updated_list[i + 1] = f"{qNos[i + 1]}B"

    return updated_list

def get_matching_value_indices(lst, value):
  return [i for i, v in enumerate(lst) if v == value]
    
def generate_first_row(worksheet,ct1, ct2, ct3):
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    worksheet.cell(row=1, column=1).value = "CLAT->"

    worksheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=3+ct1)
    worksheet.cell(row=1, column=4).value = "FT-I"

    if qpCount > 1:
        worksheet.merge_cells(start_row=1, start_column=4+ct1, end_row=1, end_column=3+ct1+ct2)
        worksheet.cell(row=1, column=4+ct1).value='FT-II'

    if qpCount > 2:
        worksheet.merge_cells(start_row=1, start_column=4+ct1+ct2, end_row=1, end_column=3+ct1+ct2+ct3)
        worksheet.cell(row=1, column=4+ct1+ct2).value='FT-III'

def generate_second_row(worksheet,coGrouping1, coGrouping2, coGrouping3):
    
    # Initialize a dictionary to store column mappings for each CO
    coColumns = {}

    # Initialize an empty list for column numbers (used temporarily)
    lstColNo = []

    # Merge the first three columns of row 2 to display the "CO ->" header
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    worksheet.cell(row=2, column=1).value = "CO ->"

    # Start assigning CO columns from column 4
    col=4
    
    # Process CO groupings for CT1
    for key,value in coGrouping1.items():
        # Merge cells corresponding to the questions for this CO
        worksheet.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col-1+len(value))
        worksheet.cell(row=2, column=col).value = "CO" + str(key)

        # Add the start and end column range for this CO to the coColumns dictionary
        coColumns[key] = [[col, col+len(value)-1]]

        # Update the column pointer for the next CO
        col+=len(value)

    # Process CO groupings for CT2, if applicable (qpCount > 1)
    if qpCount > 1:
        for key,value in coGrouping2.items():
            # Merge cells corresponding to the questions for this CO in CT2
            worksheet.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col-1+len(value))
            worksheet.cell(row=2, column=col).value = "CO" + str(key)
            
            # If this CO is already present in coColumns, retrieve its column ranges
            lstColNo = coColumns.get(key, [])
            
            # Add the new range for CT2 to the list
            lst = [col, col+len(value)-1]
            lstColNo.append(lst)

            # Update the coColumns dictionary
            coColumns[key] = lstColNo
            col+=len(value)

    # Process CO groupings for CT3, if applicable (qpCount > 2)
    if qpCount > 2:
        for key,value in coGrouping3.items():
            # Merge cells corresponding to the questions for this CO in CT3
            worksheet.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col-1+len(value))
            worksheet.cell(row=2, column=col).value = "CO" + str(key)
            # If this CO is already present in coColumns, retrieve its column ranges
            lstColNo = coColumns.get(key, [])

            # Add the new range for CT3 to the list
            lst = [col, col+len(value)-1]
            lstColNo.append(lst)

            # Update the coColumns dictionary
            coColumns[key] = lstColNo
            col+=len(value)
    return coColumns

def generate_third_row(worksheet,ct1,ct2,ct3):    

    worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)

    # Merge the first three columns in row 3 (typically for Sl.No, Register Number, and Student Name)
    worksheet.merge_cells(start_row=3, start_column=4, end_row=3, end_column=3+ct1)
    worksheet.cell(row=3, column=4).value = 'THEORY (for either/or Q, award marks for the attempted students only)'

    # If there are more than one Course Test (qpCount > 1), process CT2
    if qpCount > 1:
        worksheet.merge_cells(start_row=3, start_column=4+ct1, end_row=3, end_column=3+ct1+ct2)
        worksheet.cell(row=3, column=4+ct1).value='THEORY (for either/or Q, award marks for the attempted students only)'
    
    # If there are more than two Course Tests (qpCount > 2), process CT3
    if qpCount > 2:
        worksheet.merge_cells(start_row=3, start_column=4+ct1+ct2, end_row=3, end_column=3+ct1+ct2+ct3)
        worksheet.cell(row=3, column=4+ct1+ct2).value='THEORY (for either/or Q, award marks for the attempted students only)'
    
def generate_fourth_row(worksheet,marks1,marks2,marks3,coGrouping1,coGrouping2,coGrouping3):
    
    # Merge the first three columns of row 4 and set the header for the "MAX. MARKS" description
    worksheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    worksheet.cell(row=4, column=1).value = "MAX. MARKS (If not applicable, leave BLANK)->"

    # Generate maximum marks for CT1 based on its CO grouping and question marks
    col=generate_Qno_marks(worksheet,4,4,coGrouping1, marks1)

    # If there are more than one Course Test (qpCount > 1), process CT2
    if qpCount > 1:    
        col=generate_Qno_marks(worksheet,4,col,coGrouping2, marks2)

    # If there are more than two Course Test (qpCount > 2), process CT3
    if qpCount > 2:    
        col=generate_Qno_marks(worksheet,4,col,coGrouping3, marks3)

def generate_Qno_marks(worksheet,row, col, coGrouping, list):
    
    # Create a flattened list of question indices from the CO grouping values
    indexList = []
    for value in coGrouping.values():
        indexList.extend(value)  

    # Iterate through the question list and write headers to the worksheet
    for i in range(0, len(list)):

        # Set column width for better visibility
        worksheet.column_dimensions[get_column_letter(col+i)].width=6

        # Write the question number or description at the specified row and column
        worksheet.cell(row=row, column=col+i).value = list[indexList[i]]
    
    
    return col+len(list)

def generate_fifth_row(worksheet,ct1,ct2,ct3):

    # Merge the first three columns in row 5 (sl.no, register number, and student name are grouped here)
    worksheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=3)

    # Merge cells for CT1 columns and set the header "Question numbers mapping"
    worksheet.merge_cells(start_row=5, start_column=4, end_row=5, end_column=3+ct1)
    worksheet.cell(row=5, column=4).value = 'Question numbers mapping'

    # Merge cells for CT2 columns and set the header if there are more than 1 course test (qpCount > 1)
    if qpCount > 1:
        worksheet.merge_cells(start_row=5, start_column=4+ct1, end_row=5, end_column=3+ct1+ct2)
        worksheet.cell(row=5, column=4+ct1).value = 'Question numbers mapping'

    # Merge cells for CT3 columns and set the header if there are more than 2 course test (qpCount > 2)
    if qpCount > 2:
        worksheet.merge_cells(start_row=5, start_column=4+ct1+ct2, end_row=5, end_column=3+ct1+ct2+ct3)
        worksheet.cell(row=5, column=4+ct1+ct2).value = 'Question numbers mapping'

def generate_sixth_row(worksheet,question_numbers1,question_numbers2,question_numbers3,coGrouping1, coGrouping2, coGrouping3):
    
    # Add header for "Sl.No" in column 1 and set its width
    worksheet.cell(row=6,column=1).value="Sl.No"
    worksheet.column_dimensions[get_column_letter(1)].width=6

    # Add header for "Register Number" in column 2 and set its width
    worksheet.cell(row=6,column=2).value="Register Number"
    worksheet.column_dimensions[get_column_letter(2)].width=20

    # Add header for "Student Name" in column 3 and set its width
    worksheet.cell(row=6,column=3).value="Student Name"
    worksheet.column_dimensions[get_column_letter(3)].width=40

    # Generate headers for question numbers and marks for CT1 and get the next column index
    col=generate_Qno_marks(worksheet,6,4,coGrouping1, question_numbers1)

    # If more than 1 CT (qpCount > 1), generate headers for CT2 and update the column index
    if qpCount > 1:    
        col=generate_Qno_marks(worksheet,6,col,coGrouping2, question_numbers2)

    # If more than 2 CTs (qpCount > 2), generate headers for CT3 and update the column index
    if qpCount > 2:    
        col=generate_Qno_marks(worksheet,6,col,coGrouping3, question_numbers3)

def generate_Formulas(worksheet,ct1,ct2,ct3, coColumns):
    
    # Generate the number of students who attempted the exam for each column set (CT1, CT2, CT3)
    generate_Rowwise_Formula(worksheet,71,"Number of Students Attempted","=COUNTA({0}7:{0}70)",ct1,ct2,ct3)

    # Generate the number of students who scored more than 65% of marks for each column set
    generate_Rowwise_Formula(worksheet,72,"Number of students who got more than 65% of marks","=COUNTIF({0}7:{0}70,\">=\"&0.65*{0}4)",ct1,ct2,ct3)
    
     # Generate the average percentage of students who scored more than 65% across multiple columns (CT-wise)
    generate_Rowwise_Formula(worksheet,73,"Percentage of students who got more than 65% of marks","=IF({0}71>0,ROUND({0}72/{0}71*100,2),\"-\")",ct1,ct2,ct3)
    
    # Generate the Course Outcome (CO) attainment level based on predefined thresholds (>=85: 3, >=75: 2, >=65: 1, <65: 0)
    generate_CTwise_Formula(worksheet,74,"Average Percentage of students who got more than 65% of marks","=IFERROR(ROUND(SUMPRODUCT({0}73:{1}73,{0}4:{1}4)/SUM({0}4:{1}4), 2),\"-\")",ct1,ct2,ct3)
    
    # Generate the Course Outcome (CO) attainment level based on predefined thresholds (>=85: 3, >=75: 2, >=65: 1, <65: 0)
    generate_CTwise_Formula(worksheet,75," CO Attainment Level (>=85:3,>=75:2,>=65:1,<65:0)","=IF({0}74>=85,3,IF({0}74>=75,2,IF({0}74>=65,1,0)))",ct1,ct2,ct3)
    
     # Generate formulas for Course Outcomes (COs) across specific columns
    generate_COWise_Formulas(worksheet,coColumns)

def generate_Rowwise_Formula(worksheet,row, text, formula,ct1,ct2,ct3):   
   
    # Merge cells in the first 3 columns and add the provided text 
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    worksheet.cell(row, column=1).value = text

    # Initialize starting column for the first set of columns (CT1)
    col=4
    
    # Process CT1: Apply formulas to individual cells in this set of columns
    for i in range(0,ct1):
        colLetter=get_column_letter(col+i)  # Get the column letter
        worksheet.cell(row,col+i).value=formula.format(colLetter)   # Apply the formula
    
    # Process CT2 if qpCount > 1
    if qpCount > 1:
        col=col+ct1 # Move to the starting column for CT2
        for i in range(0,ct2):
            colLetter=get_column_letter(col+i) # Get the column letter
            worksheet.cell(row,col+i).value=formula.format(colLetter) # Apply the formula

    # Process CT3 if qpCount > 2
    if qpCount > 2:
        col=col+ct2 # Move to the starting column for CT3
        for i in range(0,ct3):
            colLetter=get_column_letter(col+i) # Get the column letter
            worksheet.cell(row,column=col+i).value=formula.format(colLetter) # Apply the formula

def generate_CTwise_Formula(worksheet,row,text,formula,ct1,ct2,ct3):    
    
    # Merge cells for the text column and add the text
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    worksheet.cell(row, column=1).value = text

     # Initialize the starting column for CT formulas
    col=4

     # Handle the formula and cell merging for CT1
    worksheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+ct1-1)
    worksheet.cell(row,col).value=formula.format(get_column_letter(col), get_column_letter(col+ct1-1))

    # Process additional CTs if qpCount > 1 or qpCount > 2
    if qpCount > 1:
        col=col+ct1  # Move to the next set of columns
        worksheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+ct2-1)
        worksheet.cell(row,col).value=formula.format(get_column_letter(col), get_column_letter(col+ct2-1))

    if qpCount > 2:
        col=col+ct2  # Move to the next set of columns
        worksheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+ct3-1)
        worksheet.cell(row,col).value=formula.format(get_column_letter(col), get_column_letter(col+ct3-1))

def generate_COWise_Formulas(worksheet, coColumns):
    
    # Generate the CO-wise table header starting at row 77, with relevant titles
    generate_CO_wise_table(worksheet,77,"CO","CO Wise Average Percentage of students who got more than 65% of marks","Overall CO Attainment Level (>=85:3,>=75:2,>=65:1,<65:0)",True)
    
    # Starting row for CO formulas
    row = 78
    
    for i in range(1,7):
        if i not in coColumns:
            # If a CO is missing, generate a row indicating "Not Applicable"
            generate_CO_wise_table(worksheet, row, f"CO{i}", 0, 0, False)
        else:
            key = i
            value = coColumns[i]        
            text1 = "CO" + str(key)# CO label, like CO1, CO2, etc.

            # Base of the formula for the CO-wise average calculation
            text2 = "=IFERROR(ROUND(("

            # Define the sub-formula structure for each pair of columns (like SUMPRODUCT)
            sub_formula = "SUMPRODUCT({0}73:{1}73,{0}4:{1}4)/SUM({0}4:{1}4)"
            num_subformulas = len(value) # Number of column pairs for the current CO

            # Iterate through the column pairs and append them to the formula   
            for i, col_pair in enumerate(value):
                start_col_letter = get_column_letter(col_pair[0])# Convert column number to letter
                end_col_letter = get_column_letter(col_pair[1])# Convert column number to letter

                # Append sub-formula for this column pair
                text2 += sub_formula.format(start_col_letter, end_col_letter)

                # Add '+' if not the last pair, else complete the formula structure
                if i != num_subformulas - 1:
                    text2 += "+"
                else:
                    text2 += f")/{num_subformulas},2),\"-\")"
            
            # Add the attainment level formula based on the CO value
            text3 = f"=IF(C{row}>=85,3,IF(C{row}>=75,2,IF(C{row}>=65,1,0)))"

            # Generate the row with the CO-wise formula and the attainment level formula
            generate_CO_wise_table(worksheet, row, text1, text2, text3, False)

        # Move to the next row for the next CO formula
        row += 1

def generate_CO_wise_table(worksheet,row,text1,text2,text3,header):

    grey_fill = styles.PatternFill(start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")

    # Merge cells for the first column (CO label) and set its value
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    worksheet.cell(row, column=1).value = text1
    worksheet.cell(row, column=1).alignment =styles.Alignment(horizontal='center', vertical='center')

    # Merge cells for the second column (CO-wise formula) and set its value
    worksheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
    worksheet.cell(row, column=3).value = text2
    worksheet.cell(row, column=3).alignment =styles.Alignment(horizontal='center', vertical='center')

    # Merge cells for the third column (attainment level formula) and set its value
    worksheet.merge_cells(start_row=row, start_column=10, end_row=row, end_column=18)
    worksheet.cell(row, column=10).value = text3
    worksheet.cell(row, column=10).alignment =styles.Alignment(horizontal='center', vertical='center')

    # Apply bold font if this is a header row, otherwise apply regular font
    if header:
        worksheet.cell(row, column=1).font = font_bold
        worksheet.cell(row, column=3).font = font_bold
        worksheet.cell(row, column=10).font = font_bold
        worksheet.cell(row, column=1).fill = grey_fill
        worksheet.cell(row, column=3).fill = grey_fill
        worksheet.cell(row, column=10).fill = grey_fill
    else:
        worksheet.cell(row, column=1).font = font
        worksheet.cell(row, column=3).font = font
        worksheet.cell(row, column=10).font = font
        # Apply bold border to all the cells in the merged ranges
    
    for col in range(1, 3):
        worksheet.cell(row, column=col).border = bold_border  # First column merged cells
    for col in range(3, 10):
        worksheet.cell(row, column=col).border = bold_border  # Second column merged cells
    for col in range(10, 19):
        worksheet.cell(row, column=col).border = bold_border  # Third column merged cells

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def server_error(e):
    return render_template('500.html'), 500

if __name__ == '__main__':
    app.run(debug=True)