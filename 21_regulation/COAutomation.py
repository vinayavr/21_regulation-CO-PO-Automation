from flask import Flask, render_template, make_response,request, jsonify, send_file
from openpyxl import Workbook, styles,load_workbook
from openpyxl.styles import Side,PatternFill
import pandas as pd
import os
import io
from io import BytesIO
import os.path
from flask import send_file
from openpyxl.styles import Font,Border
from openpyxl.utils import get_column_letter
import pdfplumber
import re
from openpyxl import workbook
from FINALTEST import second_bp


# Define a bold border  
bold_border = styles.Border(left=styles.Side(border_style='thin', color='000000'),
                right=styles.Side(border_style='thin', color='000000'),
                top=styles.Side(border_style='thin', color='000000'),
                bottom=styles.Side(border_style='thin', color='000000'))

# set font
font_bold = Font(name="Times New Roman", size=10, bold=True)
font = Font(name="Times New Roman", size=10)

qpCount = 0
coCount = 6
maxRows = 71

app = Flask(__name__)
app.register_blueprint(second_bp)

app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'pdf'}

# Function to check if the file type is allowed
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload1', methods=['POST'])
def upload():
    if 'pdf_files' not in request.files:
        return jsonify({"message": "No files part in the request"}), 400

    uploaded_files = request.files.getlist('pdf_files')
    if not uploaded_files or all(file.filename == '' for file in uploaded_files):
        return jsonify({"message": "No selected files"}), 400

    upload_folder = 'uploads'
    os.makedirs(upload_folder, exist_ok=True)

    saved_files = []
    for file in uploaded_files:
        if file and file.filename.endswith('.pdf'):
            file_path = os.path.join(upload_folder, file.filename)
            file.save(file_path)
            saved_files.append(file_path)

    # Pass the saved files to the generate_excel function
    excel_file_path = generate_excel(saved_files)

    '''return jsonify({
        "message": f"Successfully processed {len(saved_files)} file(s).",
        "generated_excel": excel_file_path
    }), 200'''
    return jsonify({
    "success": True,
    "message": f"Successfully processed {len(saved_files)} file(s).",
    "download_url": "/download/result.xlsx"  # Adjust the download URL as needed
}), 200




@app.route("/generate_excel", methods=["POST"])
def generate_excel(pdf_paths):
    
    # Initialize empty lists for storing data for each set of questions, marks, and COs
    ct1_QNos = []
    ct1_Marks = []
    ct1_COs = []
    ct2_QNos = []
    ct2_Marks = []
    ct2_COs = []
    ct3_QNos = []
    ct3_Marks = []
    ct3_COs = []

    # Determine the number of question papers based on the number of pdf_paths
    ct1_grouping = {}
    ct2_grouping = {}
    ct3_grouping = {}
    
    # Determine the number of question papers based on the number of pdf_paths
    global qpCount
    qpCount = len(pdf_paths)

    # Extract details from the first question paper PDF
    ct1_grouping = extract_details_from_pdf(pdf_paths[0], ct1_QNos, ct1_Marks, ct1_COs)
    
    # Update the Question No Choices
    ct1_QNos = update_QuestionNo_Choices(ct1_QNos)

    # If there are more than one question papers, process the second one
    if qpCount > 1:
        ct2_grouping = extract_details_from_pdf(pdf_paths[1], ct2_QNos, ct2_Marks, ct2_COs)
        # Update the Question No Choices
        ct2_QNos = update_QuestionNo_Choices(ct2_QNos)

    # If there are more than two question papers, process the third one 
    if qpCount > 2:
        ct3_grouping = extract_details_from_pdf(pdf_paths[2], ct3_QNos, ct3_Marks, ct3_COs)
        # Update the Question No Choices
        ct3_QNos = update_QuestionNo_Choices(ct3_QNos)

    # Create a new workbook
    workbook = Workbook()
    worksheet = workbook.active

    # Calculate the length of question numbers for each question paper  
    ct1=len(ct1_QNos)
    ct2=len(ct2_QNos)
    ct3=len(ct3_QNos)

    # Generate various rows in the worksheet (headers, data, and formulas)
    generate_first_row(worksheet,ct1,ct2,ct3)
    coColumns = generate_second_row(worksheet,ct1_grouping,ct2_grouping,ct3_grouping)
    generate_third_row(worksheet,ct1,ct2,ct3)
    generate_fourth_row(worksheet,ct1_Marks,ct2_Marks,ct3_Marks,ct1_grouping,ct2_grouping,ct3_grouping)
    generate_fifth_row(worksheet,ct1,ct2,ct3)
    generate_sixth_row(worksheet,ct1_QNos, ct2_QNos, ct3_QNos,ct1_grouping, ct2_grouping, ct3_grouping)
    generate_Formulas(worksheet,ct1,ct2,ct3, coColumns)
    apply_styles(worksheet)

    # Define the path for the output folder and file
    '''output_folder = os.path.join(os.getcwd(), "output")
    os.makedirs(output_folder, exist_ok=True)
    file_path = os.path.join(output_folder, "result.xlsx")

    # Save the workbook
    workbook.save(file_path)
    return file_path'''
    # Define the static folder for downloads
    static_folder = os.path.join(os.getcwd(), "static")
    os.makedirs(static_folder, exist_ok=True)  # Ensure the folder exists

    # Save file to static folder
    file_path = os.path.join(static_folder, "result.xlsx")

    # Save to disk before sending
    workbook.save(file_path)  

    # Send the file as a download
    return send_file(file_path, as_attachment=True, download_name="result.xlsx",
                 mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
def apply_styles(worksheet):
    # Define colors
    light_green_fill = styles.PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    orange_fill = styles.PatternFill(start_color="FFC300", end_color="FFC300", fill_type="solid")
    yellow_fill = styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    grey_fill = styles.PatternFill(start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")

    # Apply formatting (borders, fills, font styles, and alignment)
    for row_num,row in enumerate(worksheet.iter_rows(min_row=1, max_row=maxRows), 1):
            for cell in row:
                cell.border=bold_border  # Apply border to all cells
            # Apply conditional fill color based on row number
                match row_num:
                    case 1:
                        cell.fill = light_green_fill
                    case 2:
                        cell.fill = orange_fill
                    case 4:
                        cell.fill = yellow_fill
                    case 5:
                        if cell.column != 1:
                            cell.fill = grey_fill
                    case 68:
                        cell.fill = yellow_fill
                # Bold font for certain rows
                if row_num <= 6 or (row_num >= 67 and cell.column < 4):
                    cell.font = font_bold
                else:
                    cell.font = font
                    
                # Center alignment for all columns except the first
                if cell.column != 3:
                    cell.alignment = styles.Alignment(horizontal='center', vertical='center')

def extract_details_from_pdf(pdf_path, question_numbers, marks, co_lists):
    """
    Extracts question numbers (Q.no) from a question paper PDF.

    Args:
        pdf_path (str): Path to the PDF file.

    Returns:
        list: A list of extracted question numbers as integers.
    """

    coGrouping = {}

    # Define a regex pattern to match question numbers (e.g., "1", "2", etc.)
    flag=0

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            # Extract all text from the page
            text = page.extract_text()

            if text: 
                text = text.replace('&', '"&"')
                 # Ensure the page has text
                # Split the text into lines
                lines = text.split("\n")

                for line in lines:
                    if flag==0 and line.find("Part")!=-1:
                        flag=1
                        continue
                    if flag==1:
                        match = re.search(r".\d \d \d \d$",line.strip()) 
                        question_no= re.match(r"^(\d{2})|\d",line.strip()) 
                        if question_no and match:
                            # Add the matched number to the list
                            try:
                                qnum=question_no.group().strip()
                                question_numbers.append("Q"+qnum)
                                marks.append(int(match.group().strip().split()[0]))
                                co_lists.append(match.group().strip().split()[2])

                            except ValueError:
                                pass  # Skip if the match isn't a valid integer

    # Get the CO grouping indices
    for i in range(1,coCount+1):
        co = get_matching_value_indices(co_lists,str(i))
        if len(co) > 0:
            coGrouping[i] = co
    return coGrouping

def update_QuestionNo_Choices(qNos):
    """
    Finds two consecutive indices with the same value in a list,
    appends 'A' to the first and 'B' to the second, and returns the updated list.
    """
    updated_list = qNos.copy()  # Work with a copy of the list to avoid modifying the original
    
    for i in range(len(qNos) - 1):
        if qNos[i] == qNos[i + 1]:
            updated_list[i] = f"{qNos[i]}A"
            updated_list[i + 1] = f"{qNos[i + 1]}B"

    return updated_list


def get_matching_value_indices(lst, value):
  """
  Returns a list of indices where the specified value occurs in the given list.

  Args:
    lst: The input list.
    value: The value to search for.

  Returns:
    A list of indices where the value occurs.
  """
  return [i for i, v in enumerate(lst) if v == value]
    
def generate_first_row(worksheet,ct1, ct2, ct3):
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    worksheet.cell(row=1, column=1).value = "CLAT->"

    worksheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=3+ct1)
    worksheet.cell(row=1, column=4).value = "FT-I"

    if qpCount > 1:
        worksheet.merge_cells(start_row=1, start_column=4+ct1, end_row=1, end_column=3+ct1+ct2)
        worksheet.cell(row=1, column=4+ct1).value='FT-II'

    if qpCount > 2:
        worksheet.merge_cells(start_row=1, start_column=4+ct1+ct2, end_row=1, end_column=3+ct1+ct2+ct3)
        worksheet.cell(row=1, column=4+ct1+ct2).value='FT-III'

def generate_second_row(worksheet,coGrouping1, coGrouping2, coGrouping3):

    """
    Generates the second row of the worksheet to map Course Outcomes (COs)
    to the corresponding question columns for each Course Test (CT).

    Args:
        worksheet: The Excel worksheet object.
        coGrouping1, coGrouping2, coGrouping3: Dictionaries representing the CO groupings 
                                               for CT1, CT2, and CT3. Each key is a CO number, 
                                               and the value is a list of question indices.

    Returns:
        coColumns: A dictionary where each CO is mapped to a list of start and end column ranges 
                   for all CTs (CT1, CT2, CT3).
    """
    
    # Initialize a dictionary to store column mappings for each CO
    coColumns = {}

    # Initialize an empty list for column numbers (used temporarily)
    lstColNo = []

    # Merge the first three columns of row 2 to display the "CO ->" header
    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    worksheet.cell(row=2, column=1).value = "CO ->"

    # Start assigning CO columns from column 4
    col=4
    
    # Process CO groupings for CT1
    for key,value in coGrouping1.items():
        # Merge cells corresponding to the questions for this CO
        worksheet.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col-1+len(value))
        worksheet.cell(row=2, column=col).value = "CO" + str(key)

        # Add the start and end column range for this CO to the coColumns dictionary
        coColumns[key] = [[col, col+len(value)-1]]

        # Update the column pointer for the next CO
        col+=len(value)

    # Process CO groupings for CT2, if applicable (qpCount > 1)
    if qpCount > 1:
        for key,value in coGrouping2.items():
            # Merge cells corresponding to the questions for this CO in CT2
            worksheet.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col-1+len(value))
            worksheet.cell(row=2, column=col).value = "CO" + str(key)
            
            # If this CO is already present in coColumns, retrieve its column ranges
            lstColNo = coColumns.get(key, [])
            
            # Add the new range for CT2 to the list
            lst = [col, col+len(value)-1]
            lstColNo.append(lst)

            # Update the coColumns dictionary
            coColumns[key] = lstColNo
            col+=len(value)


    # Process CO groupings for CT3, if applicable (qpCount > 2)
    if qpCount > 2:
        for key,value in coGrouping3.items():
            # Merge cells corresponding to the questions for this CO in CT3
            worksheet.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col-1+len(value))
            worksheet.cell(row=2, column=col).value = "CO" + str(key)
            # If this CO is already present in coColumns, retrieve its column ranges
            lstColNo = coColumns.get(key, [])

            # Add the new range for CT3 to the list
            lst = [col, col+len(value)-1]
            lstColNo.append(lst)

            # Update the coColumns dictionary
            coColumns[key] = lstColNo
            col+=len(value)
    return coColumns

def generate_third_row(worksheet,ct1,ct2,ct3):    

    """
    Generates the third row of the worksheet to display the "THEORY" header
    for question categories, organized by the number of Course Tests (CTs).

    Args:
        worksheet: The Excel worksheet object.
        ct1, ct2, ct3: Column counts allocated for CT1, CT2, and CT3.

    Returns:
        None
    """
    worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)

    # Merge the first three columns in row 3 (typically for Sl.No, Register Number, and Student Name)
    worksheet.merge_cells(start_row=3, start_column=4, end_row=3, end_column=3+ct1)
    worksheet.cell(row=3, column=4).value = 'THEORY (for either/or Q, award marks for the attempted students only)'

    # If there are more than one Course Test (qpCount > 1), process CT2
    if qpCount > 1:
        worksheet.merge_cells(start_row=3, start_column=4+ct1, end_row=3, end_column=3+ct1+ct2)
        worksheet.cell(row=3, column=4+ct1).value='THEORY (for either/or Q, award marks for the attempted students only)'
    
    # If there are more than two Course Tests (qpCount > 2), process CT3
    if qpCount > 2:
        worksheet.merge_cells(start_row=3, start_column=4+ct1+ct2, end_row=3, end_column=3+ct1+ct2+ct3)
        worksheet.cell(row=3, column=4+ct1+ct2).value='THEORY (for either/or Q, award marks for the attempted students only)'
    
def generate_fourth_row(worksheet,marks1,marks2,marks3,coGrouping1,coGrouping2,coGrouping3):

    """
    Generates the fourth row of the worksheet to display the maximum marks 
    for each question, organized by Course Test (CT) and Course Outcome (CO) groupings.

    Args:
        worksheet: The Excel worksheet object.
        marks1, marks2, marks3: Lists containing the maximum marks for each question 
                                in CT1, CT2, and CT3 respectively.
        coGrouping1, coGrouping2, coGrouping3: CO groupings for CT1, CT2, and CT3 respectively.

    Returns:
        None
    """
    
    # Merge the first three columns of row 4 and set the header for the "MAX. MARKS" description
    worksheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    worksheet.cell(row=4, column=1).value = "MAX. MARKS (If not applicable, leave BLANK)->"

    # Generate maximum marks for CT1 based on its CO grouping and question marks
    col=generate_Qno_marks(worksheet,4,4,coGrouping1, marks1)

    # If there are more than one Course Test (qpCount > 1), process CT2
    if qpCount > 1:    
        col=generate_Qno_marks(worksheet,4,col,coGrouping2, marks2)

    # If there are more than two Course Test (qpCount > 2), process CT3
    if qpCount > 2:    
        col=generate_Qno_marks(worksheet,4,col,coGrouping3, marks3)

def generate_Qno_marks(worksheet,row, col, coGrouping, list):
    """
    Dynamically generates question number headers and adjusts column widths 
    for the given row in the worksheet.

    Args:
        worksheet: The Excel worksheet object.
        row: The row number where the headers will be written.
        col: The starting column number where the question headers begin.
        coGrouping: A dictionary mapping CO groupings to question indices.
        list: A list of question numbers or descriptions.

    Returns:
        The next available column index after writing the question headers.
    """
    
    # Create a flattened list of question indices from the CO grouping values
    indexList = []
    for value in coGrouping.values():
        indexList.extend(value)  

    # Iterate through the question list and write headers to the worksheet
    for i in range(0, len(list)):

        # Set column width for better visibility
        worksheet.column_dimensions[get_column_letter(col+i)].width=6

        # Write the question number or description at the specified row and column
        worksheet.cell(row=row, column=col+i).value = list[indexList[i]]
    
    # Return the next available column index after processing all questions
    return col+len(list)

def generate_fifth_row(worksheet,ct1,ct2,ct3):

    """
    Generates the fifth row of the worksheet to display "Question numbers mapping" headers 
    for each Course Test (CT) based on the number of columns allocated for CT1, CT2, and CT3.

    Args:
        worksheet: The Excel worksheet object.
        ct1, ct2, ct3: Column counts allocated for CT1, CT2, and CT3.
    """

    # Merge the first three columns in row 5 (sl.no, register number, and student name are grouped here)
    worksheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=3)

    # Merge cells for CT1 columns and set the header "Question numbers mapping"
    worksheet.merge_cells(start_row=5, start_column=4, end_row=5, end_column=3+ct1)
    worksheet.cell(row=5, column=4).value = 'Question numbers mapping'

    # Merge cells for CT2 columns and set the header if there are more than 1 course test (qpCount > 1)
    if qpCount > 1:
        worksheet.merge_cells(start_row=5, start_column=4+ct1, end_row=5, end_column=3+ct1+ct2)
        worksheet.cell(row=5, column=4+ct1).value = 'Question numbers mapping'

    # Merge cells for CT3 columns and set the header if there are more than 2 course test (qpCount > 2)
    if qpCount > 2:
        worksheet.merge_cells(start_row=5, start_column=4+ct1+ct2, end_row=5, end_column=3+ct1+ct2+ct3)
        worksheet.cell(row=5, column=4+ct1+ct2).value = 'Question numbers mapping'

def generate_sixth_row(worksheet,question_numbers1,question_numbers2,question_numbers3,coGrouping1, coGrouping2, coGrouping3):
    
    """
    Generates the sixth row of the worksheet with headers for student information and question numbers,
    and dynamically assigns columns for question marks based on Course Outcome (CO) groupings.

    Args:
        worksheet: The Excel worksheet object.
        question_numbers1, question_numbers2, question_numbers3: Lists of question numbers for CT1, CT2, and CT3.
        coGrouping1, coGrouping2, coGrouping3: CO groupings corresponding to CT1, CT2, and CT3.
    """
    # Add header for "Sl.No" in column 1 and set its width
    worksheet.cell(row=6,column=1).value="Sl.No"
    worksheet.column_dimensions[get_column_letter(1)].width=6

    # Add header for "Register Number" in column 2 and set its width
    worksheet.cell(row=6,column=2).value="Register Number"
    worksheet.column_dimensions[get_column_letter(2)].width=20

    # Add header for "Student Name" in column 3 and set its width
    worksheet.cell(row=6,column=3).value="Student Name"
    worksheet.column_dimensions[get_column_letter(3)].width=40

    # Generate headers for question numbers and marks for CT1 and get the next column index
    col=generate_Qno_marks(worksheet,6,4,coGrouping1, question_numbers1)

    # If more than 1 CT (qpCount > 1), generate headers for CT2 and update the column index
    if qpCount > 1:    
        col=generate_Qno_marks(worksheet,6,col,coGrouping2, question_numbers2)

    # If more than 2 CTs (qpCount > 2), generate headers for CT3 and update the column index
    if qpCount > 2:    
        col=generate_Qno_marks(worksheet,6,col,coGrouping3, question_numbers3)

def generate_Formulas(worksheet,ct1,ct2,ct3, coColumns):
    """
    Generates various formulas for a worksheet to calculate and analyze student performance.

    Args:
        worksheet: The Excel worksheet object.
        ct1, ct2, ct3: Column counts for CT1, CT2, and CT3 respectively.
        coColumns: List or range of columns associated with Course Outcomes (COs).
    """
    # Generate the number of students who attempted the exam for each column set (CT1, CT2, CT3)
    generate_Rowwise_Formula(worksheet,67,"Number of Students Attempted","=COUNTA({0}7:{0}66)",ct1,ct2,ct3)

    # Generate the number of students who scored more than 65% of marks for each column set
    generate_Rowwise_Formula(worksheet,68,"Number of students who got more than 65% of marks","=COUNTIF({0}7:{0}66,\">=\"&0.65*{0}4)",ct1,ct2,ct3)
    
     # Generate the average percentage of students who scored more than 65% across multiple columns (CT-wise)
    generate_Rowwise_Formula(worksheet,69,"Percentage of students who got more than 65% of marks","=IF({0}67>0,{0}68/{0}67*100,\"-\")",ct1,ct2,ct3)
    
    # Generate the Course Outcome (CO) attainment level based on predefined thresholds (>=85: 3, >=75: 2, >=65: 1, <65: 0)
    generate_CTwise_Formula(worksheet,70,"Average Percentage of students who got more than 65% of marks","=IFERROR(ROUND(SUMPRODUCT({0}69:{1}69,{0}4:{1}4)/SUM({0}4:{1}4), 2),\"-\")",ct1,ct2,ct3)
    
    # Generate the Course Outcome (CO) attainment level based on predefined thresholds (>=85: 3, >=75: 2, >=65: 1, <65: 0)
    generate_CTwise_Formula(worksheet,71," CO Attainment Level (>=85:3,>=75:2,>=65:1,<65:0)","=IF({0}70>=85,3,IF({0}70>=75,2,IF({0}70>=65,1,0)))",ct1,ct2,ct3)
    
     # Generate formulas for Course Outcomes (COs) across specific columns
    generate_COWise_Formulas(worksheet,coColumns)

def generate_Rowwise_Formula(worksheet,row, text, formula,ct1,ct2,ct3):   
    """
    Generates and populates a worksheet row with formulas applied to individual cells, customized for row-wise logic.

    Args:
        worksheet: The Excel worksheet object.
        row: Row number to populate.
        text: Text to be placed in the first set of merged cells.
        formula: A format string for the formula to be applied to individual cells.
        ct1, ct2, ct3: Column counts for CT1, CT2, and CT3 respectively.
        qpCount: Number of CTs to include (1, 2, or 3).
    """
    # Merge cells in the first 3 columns and add the provided text 
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    worksheet.cell(row, column=1).value = text

    # Initialize starting column for the first set of columns (CT1)
    col=4
    
    # Process CT1: Apply formulas to individual cells in this set of columns
    for i in range(0,ct1):
        colLetter=get_column_letter(col+i)  # Get the column letter
        worksheet.cell(row,col+i).value=formula.format(colLetter)   # Apply the formula
    
    # Process CT2 if qpCount > 1
    if qpCount > 1:
        col=col+ct1 # Move to the starting column for CT2
        for i in range(0,ct2):
            colLetter=get_column_letter(col+i) # Get the column letter
            worksheet.cell(row,col+i).value=formula.format(colLetter) # Apply the formula

    # Process CT3 if qpCount > 2
    if qpCount > 2:
        col=col+ct2 # Move to the starting column for CT3
        for i in range(0,ct3):
            colLetter=get_column_letter(col+i) # Get the column letter
            worksheet.cell(row,column=col+i).value=formula.format(colLetter) # Apply the formula

def generate_CTwise_Formula(worksheet,row,text,formula,ct1,ct2,ct3):    
    """
    Generates and populates a worksheet row with merged cells and formulas, customized for column-wise (CT-wise) logic.

    Args:
        worksheet: The Excel worksheet object.
        row: Row number to populate.
        text: Text to be placed in the first set of merged cells.
        formula: A format string for the formula to be applied to merged cells.
        ct1, ct2, ct3: Column counts for CT1, CT2, and CT3 respectively.
        qpCount: Number of CTs to include (1, 2, or 3).
    """
    # Merge cells for the text column and add the text
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    worksheet.cell(row, column=1).value = text

     # Initialize the starting column for CT formulas
    col=4

     # Handle the formula and cell merging for CT1
    worksheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+ct1-1)
    worksheet.cell(row,col).value=formula.format(get_column_letter(col), get_column_letter(col+ct1-1))

    # Process additional CTs if qpCount > 1 or qpCount > 2
    if qpCount > 1:
        col=col+ct1  # Move to the next set of columns
        worksheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+ct2-1)
        worksheet.cell(row,col).value=formula.format(get_column_letter(col), get_column_letter(col+ct2-1))

    if qpCount > 2:
        col=col+ct2  # Move to the next set of columns
        worksheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+ct3-1)
        worksheet.cell(row,col).value=formula.format(get_column_letter(col), get_column_letter(col+ct3-1))

def generate_COWise_Formulas(worksheet, coColumns):
    """
    Generates CO-wise formulas in an Excel worksheet using column *numbers*.

    Args:
        worksheet: The openpyxl worksheet object.
        coColumns: A dictionary where keys are CO numbers and values are lists of column number pairs.
            Example: {1: [[1, 2], [3, 4]], 2: [[5, 6]]} (1=A, 2=B, etc.)
    """
    
    # Generate the CO-wise table header starting at row 73, with relevant titles
    generate_CO_wise_table(worksheet,73,"CO","CO Wise Average Percentage of students who got more than 65% of marks","Overall CO Attainment Level (>=85:3,>=75:2,>=65:1,<65:0)",True)
    
    # Starting row for CO formulas
    row = 74    
    
    # Iterate over the coColumns dictionary to generate formulas for each CO
    for key, value in coColumns.items():
        text1 = "CO" + str(key)# CO label, like CO1, CO2, etc.

        # Base of the formula for the CO-wise average calculation
        text2 = "=IFERROR(ROUND(("

        # Define the sub-formula structure for each pair of columns (like SUMPRODUCT)
        sub_formula = "SUMPRODUCT({0}69:{1}69,{0}4:{1}4)/SUM({0}4:{1}4)"
        num_subformulas = len(value) # Number of column pairs for the current CO

        # Iterate through the column pairs and append them to the formula   
        for i, col_pair in enumerate(value):
            start_col_letter = get_column_letter(col_pair[0])# Convert column number to letter
            end_col_letter = get_column_letter(col_pair[1])# Convert column number to letter

            # Append sub-formula for this column pair
            text2 += sub_formula.format(start_col_letter, end_col_letter)

            # Add '+' if not the last pair, else complete the formula structure
            if i != num_subformulas - 1:
                text2 += "+"
            else:
                text2 += f")/{num_subformulas},2),\"-\")"
        
        # Add the attainment level formula based on the CO value
        text3 = f"=IF(C{row}>=85,3,IF(C{row}>=75,2,IF(C{row}>=65,1,0)))"

        # Generate the row with the CO-wise formula and the attainment level formula
        generate_CO_wise_table(worksheet, row, text1, text2, text3, False)

        # Move to the next row for the next CO formula
        row += 1

def generate_CO_wise_table(worksheet,row,text1,text2,text3,header):
    """
    Generates a row in the worksheet with CO-wise data, merging cells for better readability.

    Args:
        worksheet: The openpyxl worksheet object where the table is generated.
        row: The row number where the data should be placed.
        text1: The content for the first column (CO label).
        text2: The content for the second column (CO-wise formula).
        text3: The content for the third column (attainment level formula).
        header: Boolean value to determine if the row is a header.
            - If True, the text is bolded.
            - If False, the text uses regular font style.
    """
    grey_fill = styles.PatternFill(start_color="c0c0c0", end_color="c0c0c0", fill_type="solid")

    # Merge cells for the first column (CO label) and set its value
    worksheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    worksheet.cell(row, column=1).value = text1
    worksheet.cell(row, column=1).alignment =styles.Alignment(horizontal='center', vertical='center')

    # Merge cells for the second column (CO-wise formula) and set its value
    worksheet.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
    worksheet.cell(row, column=3).value = text2
    worksheet.cell(row, column=3).alignment =styles.Alignment(horizontal='center', vertical='center')

    # Merge cells for the third column (attainment level formula) and set its value
    worksheet.merge_cells(start_row=row, start_column=11, end_row=row, end_column=20)
    worksheet.cell(row, column=11).value = text3
    worksheet.cell(row, column=11).alignment =styles.Alignment(horizontal='center', vertical='center')

    # Apply bold font if this is a header row, otherwise apply regular font
    if header:
        worksheet.cell(row, column=1).font = font_bold
        worksheet.cell(row, column=3).font = font_bold
        worksheet.cell(row, column=11).font = font_bold
        worksheet.cell(row, column=1).fill = grey_fill
        worksheet.cell(row, column=3).fill = grey_fill
        worksheet.cell(row, column=11).fill = grey_fill
    else:
        worksheet.cell(row, column=1).font = font
        worksheet.cell(row, column=3).font = font
        worksheet.cell(row, column=11).font = font
        # Apply bold border to all the cells in the merged ranges
    
    for col in range(1, 3):
        worksheet.cell(row, column=col).border = bold_border  # First column merged cells
    for col in range(3, 11):
        worksheet.cell(row, column=col).border = bold_border  # Second column merged cells
    for col in range(11, 21):
        worksheet.cell(row, column=col).border = bold_border  # Third column merged cells
    
if __name__ == '__main__':
    app.run(debug=True)




