import traceback
import os
import sys
import json
import logging
from typing import Dict, List, Optional
from copy import copy

import pdfplumber
import pandas as pd
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re

from flask import Blueprint, request, abort, send_from_directory, render_template, jsonify, send_file, current_app, Flask

second_bp = Blueprint('second', __name__)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s: %(message)s',
    handlers=[
        logging.FileHandler('app.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

class TLPMarkConverter:
    def __init__(self, config: Dict = None):
        self.config = config or {
            'upload_dir': './uploads',
            'results_dir': './static',
            'max_file_size': 50 * 1024 * 1024,
            'allowed_extensions': {'.pdf'},
            'allowed_excel_extensions': {'.xlsx', '.xls'}
        }
        self.logger = logger
        
        os.makedirs(self.config['upload_dir'], exist_ok=True)
        os.makedirs(self.config['results_dir'], exist_ok=True)

    def validate_file(self, filename: str) -> bool:
        return os.path.splitext(filename)[1].lower() in self.config['allowed_extensions']

    def validate_excel_file(self, filename: str) -> bool:
        return os.path.splitext(filename)[1].lower() in self.config['allowed_excel_extensions']

    def extract_marks_from_tlp(self, pdf_files: List[str]) -> Dict[str, Dict]:

        marks_data = {}
        stats = {
            'total_files': len(pdf_files),
            'processed_files': 0,
            'failed_files': 0,
            'total_entries': 0,
            'file_stats': {}
        }
        
        # Regular expression for conducted max value
        conducted_max_pattern = re.compile(r"Conducted Max\.?\s+(\d+\.?\d*)")
        
        # Regular expression for register numbers and marks (handles A and 0A)
        reg_pattern = re.compile(r"(RA\d{13})\s+((?:\d+\.?\d*)|(?:[A0]A))")
        
        for pdf_path in pdf_files:
            file_name = os.path.basename(pdf_path)
            entries_found = 0
            conducted_max = None
            
            try:
                with pdfplumber.open(pdf_path) as pdf:
                    file_text = ""
                    for page in pdf.pages:
                        # Extract full text and append
                        page_text = page.extract_text() or ""
                        file_text += page_text
                    
                    # Find conducted max value first
                    max_match = conducted_max_pattern.search(file_text)
                    if max_match:
                        conducted_max = float(max_match.group(1))
                        logger.info(f"Found Conducted Max value: {conducted_max} in {file_name}")
                    else:
                        logger.warning(f"Conducted Max value not found in {file_name}")
                    
                    # Find all register number and mark matches
                    matches = reg_pattern.findall(file_text)
                    
                    for reg_no, marks in matches:
                        # Process marks value
                        if marks == 'A' or marks == '0A':
                            current_marks = 0
                        else:
                            current_marks = float(marks)
                        
                        # Aggregate marks for each register number
                        if reg_no in marks_data:
                            marks_data[reg_no] += current_marks
                        else:
                            marks_data[reg_no] = current_marks
                        
                        entries_found += 1
                
                stats['processed_files'] += 1
                stats['total_entries'] += entries_found
                stats['file_stats'][file_name] = {
                    'status': 'success',
                    'entries_found': entries_found,
                    'conducted_max': conducted_max
                }
                logger.info(f"Successfully processed {file_name}: found {entries_found} entries")
                
            except Exception as e:
                stats['failed_files'] += 1
                stats['file_stats'][file_name] = {
                    'status': 'failed',
                    'error': str(e)
                }
                logger.error(f"Error processing PDF {pdf_path}: {e}", exc_info=True)
        
        logger.info(f"Total unique entries found across all files: {len(marks_data)}")
        return {
            'marks_data': marks_data,
            'stats': stats
        }

    def copy_existing_sheets(self, source_workbook, target_workbook):
        
        try:
            for sheet_name in source_workbook.sheetnames:
                source_sheet = source_workbook[sheet_name]
                target_sheet = target_workbook.create_sheet(title=sheet_name)
                
                # Copy cell values and styles safely
                for row in source_sheet.iter_rows():
                    for cell in row:
                        target_cell = target_sheet.cell(
                            row=cell.row, 
                            column=cell.column, 
                            value=cell.value
                        )
                        
                        # Copy cell formatting safely using copy()
                        try:
                            if cell.has_style:
                                if cell.font:
                                    target_cell.font = copy(cell.font)
                                if cell.border:
                                    target_cell.border = copy(cell.border)
                                if cell.fill:
                                    target_cell.fill = copy(cell.fill)
                                if cell.alignment:
                                    target_cell.alignment = copy(cell.alignment)
                                if cell.number_format:
                                    target_cell.number_format = cell.number_format
                                if cell.protection:
                                    target_cell.protection = copy(cell.protection)
                        except Exception as style_error:
                            # If style copying fails, just log and continue
                            self.logger.warning(f"Could not copy style for cell {cell.coordinate}: {style_error}")
                            continue
                
                # Copy merged cells
                try:
                    for merged_cell_range in source_sheet.merged_cells.ranges:
                        target_sheet.merge_cells(str(merged_cell_range))
                except Exception as merge_error:
                    self.logger.warning(f"Could not copy merged cells for sheet {sheet_name}: {merge_error}")
                
                # Copy column dimensions
                try:
                    for col_letter, col_dimension in source_sheet.column_dimensions.items():
                        target_sheet.column_dimensions[col_letter].width = col_dimension.width
                except Exception as col_error:
                    self.logger.warning(f"Could not copy column dimensions for sheet {sheet_name}: {col_error}")
                
                # Copy row dimensions
                try:
                    for row_num, row_dimension in source_sheet.row_dimensions.items():
                        target_sheet.row_dimensions[row_num].height = row_dimension.height
                except Exception as row_error:
                    self.logger.warning(f"Could not copy row dimensions for sheet {sheet_name}: {row_error}")
                    
            self.logger.info(f"Successfully copied {len(source_workbook.sheetnames)} sheets from source workbook")
            
        except Exception as e:
            self.logger.error(f"Error copying sheets from source workbook: {e}", exc_info=True)
            raise

    def create_excel_sheet(
        self, 
        file_path: str, 
        append_file_path: str,
        marks_data: Dict[str, float], 
        co_splits: Dict[str, int],
        processing_stats: Dict = None
    ) -> None:
        
        self.logger.info(f"Creating Excel sheet with append_file_path: {append_file_path}")
        
        workbook = Workbook()
        # Remove default sheet if it exists
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # First, create the TLP sheet (CO Mark Distribution)
        sheet = workbook.create_sheet(title="TLP Sheet")
        
        # Check if there's an uploaded Excel file to append
        has_uploaded_excel = append_file_path and append_file_path != "" and os.path.exists(append_file_path)
        
        if has_uploaded_excel:
            try:
                source_workbook = load_workbook(append_file_path)
                self.logger.info(f"Loading existing Excel file with {len(source_workbook.sheetnames)} sheets")
                
                # Copy all existing sheets from uploaded file
                self.copy_existing_sheets(source_workbook, workbook)
                
                # Close source workbook
                source_workbook.close()
                self.logger.info("Successfully copied existing sheets from uploaded Excel file")
                
            except Exception as e:
                self.logger.error(f"Error loading existing Excel file: {e}", exc_info=True)
                # Continue with just the TLP sheet
                self.logger.info("Continuing with TLP sheet only due to error in uploaded file")
        else:
            self.logger.info("No uploaded Excel file provided, creating TLP sheet only")
                
        styles = {
            'title_font': Font(bold=True, size=12, name="Times New Roman"),
            'header_font': Font(bold=True, size=10, name="Times New Roman"),
            'calculation_font': Font(size=10, name="Times New Roman"),
            'title_fill': PatternFill(start_color="92D050", fill_type="solid"),
            'header_fill': PatternFill(start_color="F0F0F0", fill_type="solid"),
            'border': Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        }
        
        # Main Sheet - Title and headers
        sheet.merge_cells('A1:I1')
        title_cell = sheet.cell(row=1, column=1, value="CO Mark Distribution")
        title_cell.font = styles['title_font']
        title_cell.fill = styles['title_fill']
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in range(1, 10):
            sheet.cell(1, column=col).border = styles['border']  # First column merged cells
        
        # Add total CO marks split up
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
        sheet.cell(row=2, column=1, value="Total CO Marks:").font = styles['header_font']
        for col in range(1, 3):
            sheet.cell(2, column=col).border = styles['border']  # First column merged cells

        total = 0
        for col, co in enumerate(["CO1", "CO2", "CO3", "CO4", "CO5", "CO6"], 4):
            val = co_splits.get(co, 0)
            total += val
            text = co_splits.get(co, '')
            sheet.cell(row=2, column=col, value=text).font = styles['header_font']
            sheet.cell(row=2, column=col).border = styles['border']
            
        sheet.cell(row=2, column=3, value=total).font = styles['header_font']

        headers = ["S.No", "Register No", "Total Marks", "CO1", "CO2", "CO3", "CO4", "CO5", "CO6"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col, value=header)
            cell.font = styles['header_font']
            cell.fill = styles['header_fill']
            cell.alignment = Alignment(horizontal='center')
            cell.border = styles['border']

        sheet.column_dimensions[get_column_letter(1)].width = 6
        sheet.column_dimensions[get_column_letter(2)].width = 20
        sheet.column_dimensions[get_column_letter(3)].width = 40

        # Sort marks data by register number
        sorted_marks = sorted(marks_data.items(), key=lambda x: x[0])
        
        # Populate data rows
        for index, (reg_no, total_marks) in enumerate(sorted_marks, 1):
            row = index + 3
            sheet.cell(row=row, column=1, value=index).border = styles['border']
            sheet.cell(row=row, column=2, value=reg_no).border = styles['border']
            sheet.cell(row=row, column=3, value=total_marks).border = styles['border']
            
            # Calculate CO marks based on total marks and percentages
            co_total = sum(co_splits.get(f'CO{i}', 0) for i in range(1, 7))
            
            for col, co in enumerate(["CO1", "CO2", "CO3", "CO4", "CO5", "CO6"], 4):
                # Calculate proportional marks for this CO
                co_value = co_splits.get(co, 0)
                mark_text = ''
                if co_value > 0:
                    if co_total > 0:
                        co_marks = (co_value / co_total) * total_marks
                    else:
                        co_marks = 0
                    mark_text = round(co_marks, 2)

                cell = sheet.cell(row=row, column=col, value=mark_text)
                cell.border = styles['border']
                cell.alignment = Alignment(horizontal='center')
                
        # Generate the number of students who attempted 
        sheet.merge_cells(start_row=64, start_column=1, end_row=64, end_column=3)
        sheet.cell(64, column=1, value="Number of Students Attempted").font = styles['header_font']
        for col in range(1, 4):
            sheet.cell(64, column=col).border = styles['border']  # First column merged cells
        
        for i in range(0, 6):
            colLetter = get_column_letter(4 + i)  # Get the column letter
            sheet.cell(64, 4 + i).value = "=COUNTA({0}4:{0}63)".format(colLetter)   # Apply the formula
            sheet.cell(64, 4 + i).font = styles['calculation_font']
            sheet.cell(64, 4 + i).border = styles['border']
            sheet.cell(64, 4 + i).alignment = Alignment(horizontal='center', vertical='center')

        # Generate the number of students who scored more than 65% of marks
        sheet.merge_cells(start_row=65, start_column=1, end_row=65, end_column=3)
        sheet.cell(65, column=1, value="Number of students who got more than 65% of marks").font = styles['header_font']
        for col in range(1, 4):
            sheet.cell(65, column=col).border = styles['border']  # First column merged cells
        
        for i in range(0, 6):
            colLetter = get_column_letter(4 + i)  # Get the column letter
            sheet.cell(65, 4 + i).value = "=COUNTIF({0}4:{0}63,\">=\"&0.65*{0}2)".format(colLetter)   # Apply the formula            
            sheet.cell(65, 4 + i).font = styles['calculation_font']
            sheet.cell(65, 4 + i).border = styles['border']
            sheet.cell(65, 4 + i).alignment = Alignment(horizontal='center', vertical='center')

        # Generate the percentage of students who scored more than 65% 
        sheet.merge_cells(start_row=66, start_column=1, end_row=66, end_column=3)
        sheet.cell(66, column=1, value="Percentage of students who got more than 65% of marks").font = styles['header_font']
        for col in range(1, 4):
            sheet.cell(66, column=col).border = styles['border']  # First column merged cells

        for i in range(0, 6):
            colLetter = get_column_letter(4 + i)  # Get the column letter
            sheet.cell(66, 4 + i).value = "=IF({0}64>0,ROUND({0}65/{0}64*100,2),\"-\")".format(colLetter)   # Apply the formula            
            sheet.cell(66, 4 + i).font = styles['calculation_font']
            sheet.cell(66, 4 + i).border = styles['border']
            sheet.cell(66, 4 + i).alignment = Alignment(horizontal='center', vertical='center')

        # Generate the Course Outcome (CO) attainment level based on predefined thresholds (>=85: 3, >=75: 2, >=65: 1, <65: 0)
        sheet.merge_cells(start_row=67, start_column=1, end_row=67, end_column=3)
        sheet.cell(67, column=1, value=" CO Attainment Level (>=85:3,>=75:2,>=65:1,<65:0)").font = styles['header_font']
        for col in range(1, 4):
            sheet.cell(67, column=col).border = styles['border']  # First column merged cells
        
        for i in range(0, 6):
            colLetter = get_column_letter(4 + i)  # Get the column letter =IF(G64>0,(IF(G66>=85,3,IF(G66>=75,2,IF(G66>=65,1,0)))),"-")
            sheet.cell(67, 4 + i).value = "=IF({0}64>0,(IF({0}66>=85,3,IF({0}66>=75,2,IF({0}66>=65,1,0)))),\"-\")".format(colLetter)
            sheet.cell(67, 4 + i).font = styles['calculation_font']
            sheet.cell(67, 4 + i).border = styles['border']
            sheet.cell(67, 4 + i).alignment = Alignment(horizontal='center', vertical='center')

        # Auto-adjust column widths
        for col in range(4, 10):
            sheet.column_dimensions[get_column_letter(col)].auto_size = True

        workbook.save(file_path)
        self.logger.info(f"Excel sheet created: {file_path}")

@second_bp.route('/upload2', methods=['POST'])
def upload_files():
    try:
        # Get uploaded Excel file (now optional)
        co_filled_excel_file = request.files.get('co_filled_excel')
        co_filled_excel_path = None
        
        logger.info(f"Excel file received: {co_filled_excel_file.filename if co_filled_excel_file else 'None'}")
        
        converter = TLPMarkConverter()
        
        # Handle Excel file upload - now optional
        if co_filled_excel_file and co_filled_excel_file.filename:
            if converter.validate_excel_file(co_filled_excel_file.filename):
                excel_filename = secure_filename(co_filled_excel_file.filename)
                co_filled_excel_path = os.path.join(converter.config['upload_dir'], excel_filename)
                co_filled_excel_file.save(co_filled_excel_path)
                logger.info(f"Excel file saved: {co_filled_excel_path}")
            else:
                return jsonify({
                    'success': False, 
                    'message': 'Invalid Excel file format. Please upload .xlsx or .xls files only.'
                }), 400
        else:
            logger.info("No Excel file uploaded - will create TLP sheet only")

        # Get uploaded PDF files
        uploaded_files = request.files.getlist('pdf_files')
        
        # Get CO splits from form
        co_splits = {}
        for i in range(1, 7):
            co_value = request.form.get('co' + str(i), '')
            if co_value == '':
                co_splits['CO' + str(i)] = 0
            else:
                co_splits['CO' + str(i)] = float(co_value)

        if not uploaded_files or uploaded_files[0].filename == '':
            return jsonify({'success': False, 'message': 'No PDF files uploaded'}), 400

        saved_files = []
        invalid_files = []
        
        # Process PDF files
        for file in uploaded_files:
            if file and file.filename:
                if converter.validate_file(file.filename):
                    filename = secure_filename(file.filename)
                    file_path = os.path.join(converter.config['upload_dir'], filename)
                    file.save(file_path)
                    saved_files.append(file_path)
                else:
                    invalid_files.append(file.filename)

        if not saved_files:
            return jsonify({
                'success': False, 
                'message': f'No valid PDFs were saved. Invalid files: {", ".join(invalid_files)}'
            }), 400

        # Extract marks from all uploaded files
        extraction_result = converter.extract_marks_from_tlp(saved_files)
        marks_data = extraction_result['marks_data']
        stats = extraction_result['stats']
        
        if not marks_data:
            return jsonify({
                'success': False, 
                'message': 'No marks data found in PDFs. Please check the file format.'
            }), 400
        
        total_conducted_max = 0.0
    
        # Access the file_stats dictionary in the returned data
        file_stats = stats.get('file_stats', {})
        
        # Iterate through each file's stats and sum up the conducted_max values
        for file_name, file_stat in file_stats.items():
            if file_stat.get('status') == 'success' and 'conducted_max' in file_stat:
                file_conducted_max = file_stat['conducted_max']
                if file_conducted_max is not None:  # Check if the value was found
                    total_conducted_max += file_conducted_max
                    logger.info(f"Added {file_conducted_max} from {file_name} to total")
                else:
                    logger.warning(f"No Conducted Max found for {file_name}")
        
        logger.info(f"Total Conducted Max across all files: {total_conducted_max}")

        co_total = sum(co_splits.values())

        if co_total != total_conducted_max and total_conducted_max > 0:
            return jsonify({
                'success': False, 
                'message': f'CO split is not proper. Please enter correct splitup. CO total: {co_total}, Conducted Max: {total_conducted_max}'
            }), 400
        
        # Create output Excel file
        output_file = os.path.join(converter.config['results_dir'], 'co_allocation.xlsx')
        converter.create_excel_sheet(output_file, co_filled_excel_path, marks_data, co_splits, stats)
        
        # Clean up uploaded files (optional - remove if you want to keep them)
        try:
            for file_path in saved_files:
                if os.path.exists(file_path):
                    os.remove(file_path)
            if co_filled_excel_path and os.path.exists(co_filled_excel_path):
                os.remove(co_filled_excel_path)
        except Exception as e:
            logger.warning(f"Error cleaning up uploaded files: {e}")
        
        # Prepare response message based on whether Excel file was uploaded
        if co_filled_excel_path:
            success_message = (
                f"Excel file created successfully with {len(marks_data)} unique entries in TLP sheet. "
                f"Processed {stats['processed_files']} out of {stats['total_files']} files. "
                f"Final Excel file contains both TLP sheet and original uploaded sheet(s)."
            )
        else:
            success_message = (
                f"Excel file created successfully with {len(marks_data)} unique entries in TLP sheet. "
                f"Processed {stats['processed_files']} out of {stats['total_files']} files. "
                f"Final Excel file contains only the TLP sheet."
            )
        
        if stats['failed_files'] > 0:
            success_message += f" {stats['failed_files']} files could not be processed."
        
        if invalid_files:
            success_message += f" Ignored {len(invalid_files)} non-PDF files."
        
        return jsonify({
            'success': True, 
            'message': success_message,
            'processed_files': stats['processed_files'],
            'failed_files': stats['failed_files'],
            'total_entries': len(marks_data),
            'download_url': '/static/co_allocation.xlsx'
        })
    
    except Exception as e:
        logger.error(f"Unexpected error: {e}", exc_info=True)
        return jsonify({'success': False, 'message': f'Unexpected error: {str(e)}'}), 500

@second_bp.route('/download/<filename>')
def download_file(filename):
    converter = TLPMarkConverter()  # Create a converter instance
    file_path = os.path.join(converter.config['results_dir'], filename)
    return send_file(file_path, as_attachment=True) if os.path.exists(file_path) else ("File not found", 404)
